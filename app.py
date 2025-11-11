from flask import Flask, render_template, request, jsonify, session, redirect, url_for
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from werkzeug.security import generate_password_hash, check_password_hash
import os
import json
from functools import wraps

app = Flask(__name__, template_folder='app/templates', static_folder='app/static')
app.secret_key = 'your-secret-key-change-in-production'

# Configuration
DATA_DIR = 'data'
USERS_FILE = os.path.join(DATA_DIR, 'users.json')

# Ensure data directory exists
os.makedirs(DATA_DIR, exist_ok=True)

# Categories for expenses
EXPENSE_CATEGORIES = [
    'Food & Dining', 'Transportation', 'Shopping', 'Entertainment',
    'Bills & Utilities', 'Healthcare', 'Education', 'Travel',
    'Groceries', 'Rent/EMI', 'Investment', 'Other'
]

# User authentication decorator
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# User management functions
def load_users():
    """Load users from JSON file"""
    if os.path.exists(USERS_FILE):
        with open(USERS_FILE, 'r') as f:
            return json.load(f)
    return {}

def save_users(users):
    """Save users to JSON file"""
    with open(USERS_FILE, 'w') as f:
        json.dump(users, f, indent=2)

def get_user_excel_file(user_id):
    """Get the Excel file path for a user"""
    return os.path.join(DATA_DIR, f'{user_id}_expense_tracker.xlsx')

def create_user_workbook(user_id):
    """Create a new Excel workbook for a user with all necessary sheets"""
    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Create Expenses sheet
    ws_expenses = wb.create_sheet('Expenses')
    headers_expenses = ['Date', 'Category', 'Description', 'Amount (₹)', 'Payment Method', 'Notes']
    ws_expenses.append(headers_expenses)
    format_header_row(ws_expenses)

    # Create Debts sheet
    ws_debts = wb.create_sheet('Debts')
    headers_debts = ['Date', 'Creditor/Debtor', 'Type', 'Amount (₹)', 'Status', 'Due Date', 'Notes']
    ws_debts.append(headers_debts)
    format_header_row(ws_debts)

    # Create Savings sheet
    ws_savings = wb.create_sheet('Savings')
    headers_savings = ['Date', 'Type', 'Description', 'Amount (₹)', 'Goal', 'Notes']
    ws_savings.append(headers_savings)
    format_header_row(ws_savings)

    # Create Budgets sheet
    ws_budgets = wb.create_sheet('Budgets')
    headers_budgets = ['Month', 'Category', 'Budget Amount (₹)', 'Spent Amount (₹)', 'Remaining (₹)']
    ws_budgets.append(headers_budgets)
    format_header_row(ws_budgets)

    # Create Summary sheet
    ws_summary = wb.create_sheet('Summary', 0)
    ws_summary['A1'] = 'Expense Tracker Summary'
    ws_summary['A1'].font = Font(size=16, bold=True)
    ws_summary['A3'] = 'Total Expenses:'
    ws_summary['A4'] = 'Total Debts:'
    ws_summary['A5'] = 'Total Savings:'
    ws_summary['A6'] = 'Net Balance:'

    excel_file = get_user_excel_file(user_id)
    wb.save(excel_file)
    return wb

def format_header_row(worksheet):
    """Format the header row of a worksheet"""
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

def get_or_create_workbook(user_id):
    """Get existing workbook or create a new one"""
    excel_file = get_user_excel_file(user_id)
    if os.path.exists(excel_file):
        return load_workbook(excel_file)
    return create_user_workbook(user_id)

def add_expense(user_id, date, category, description, amount, payment_method, notes=''):
    """Add an expense entry to the user's Excel file"""
    wb = get_or_create_workbook(user_id)
    ws = wb['Expenses']
    ws.append([date, category, description, float(amount), payment_method, notes])
    wb.save(get_user_excel_file(user_id))

def add_debt(user_id, date, creditor, debt_type, amount, status, due_date, notes=''):
    """Add a debt entry to the user's Excel file"""
    wb = get_or_create_workbook(user_id)
    ws = wb['Debts']
    ws.append([date, creditor, debt_type, float(amount), status, due_date, notes])
    wb.save(get_user_excel_file(user_id))

def add_saving(user_id, date, saving_type, description, amount, goal, notes=''):
    """Add a saving entry to the user's Excel file"""
    wb = get_or_create_workbook(user_id)
    ws = wb['Savings']
    ws.append([date, saving_type, description, float(amount), goal, notes])
    wb.save(get_user_excel_file(user_id))

def get_expenses(user_id, limit=None):
    """Get all expenses for a user"""
    wb = get_or_create_workbook(user_id)
    ws = wb['Expenses']
    expenses = []
    for row in list(ws.iter_rows(min_row=2, values_only=True)):
        if row[0]:  # Check if date exists
            expenses.append({
                'date': row[0],
                'category': row[1],
                'description': row[2],
                'amount': row[3],
                'payment_method': row[4],
                'notes': row[5] or ''
            })
    expenses.reverse()  # Show most recent first
    return expenses[:limit] if limit else expenses

def get_debts(user_id):
    """Get all debts for a user"""
    wb = get_or_create_workbook(user_id)
    ws = wb['Debts']
    debts = []
    for row in list(ws.iter_rows(min_row=2, values_only=True)):
        if row[0]:
            debts.append({
                'date': row[0],
                'creditor': row[1],
                'type': row[2],
                'amount': row[3],
                'status': row[4],
                'due_date': row[5],
                'notes': row[6] or ''
            })
    debts.reverse()
    return debts

def get_savings(user_id):
    """Get all savings for a user"""
    wb = get_or_create_workbook(user_id)
    ws = wb['Savings']
    savings = []
    for row in list(ws.iter_rows(min_row=2, values_only=True)):
        if row[0]:
            savings.append({
                'date': row[0],
                'type': row[1],
                'description': row[2],
                'amount': row[3],
                'goal': row[4],
                'notes': row[5] or ''
            })
    savings.reverse()
    return savings

def get_summary_stats(user_id):
    """Calculate summary statistics for a user"""
    wb = get_or_create_workbook(user_id)

    # Calculate total expenses
    ws_expenses = wb['Expenses']
    total_expenses = sum(row[3] for row in ws_expenses.iter_rows(min_row=2, values_only=True) if row[0] and row[3])

    # Calculate total debts
    ws_debts = wb['Debts']
    total_debts = sum(row[3] for row in ws_debts.iter_rows(min_row=2, values_only=True) if row[0] and row[3])

    # Calculate total savings
    ws_savings = wb['Savings']
    total_savings = sum(row[3] for row in ws_savings.iter_rows(min_row=2, values_only=True) if row[0] and row[3])

    # Calculate category-wise expenses
    category_expenses = {}
    for row in ws_expenses.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1] and row[3]:
            category = row[1]
            category_expenses[category] = category_expenses.get(category, 0) + row[3]

    return {
        'total_expenses': total_expenses,
        'total_debts': total_debts,
        'total_savings': total_savings,
        'net_balance': total_savings - (total_expenses + total_debts),
        'category_expenses': category_expenses
    }

# Routes
@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        data = request.json
        username = data.get('username')
        password = data.get('password')
        email = data.get('email')

        users = load_users()

        if username in users:
            return jsonify({'success': False, 'message': 'Username already exists'}), 400

        # Create new user
        users[username] = {
            'password': generate_password_hash(password),
            'email': email,
            'created_at': datetime.now().isoformat()
        }
        save_users(users)

        # Create Excel file for user
        create_user_workbook(username)

        return jsonify({'success': True, 'message': 'Registration successful'})

    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        data = request.json
        username = data.get('username')
        password = data.get('password')

        users = load_users()

        if username in users and check_password_hash(users[username]['password'], password):
            session['user_id'] = username
            return jsonify({'success': True, 'message': 'Login successful'})

        return jsonify({'success': False, 'message': 'Invalid credentials'}), 401

    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('user_id', None)
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    user_id = session['user_id']
    stats = get_summary_stats(user_id)
    recent_expenses = get_expenses(user_id, limit=5)

    return render_template('dashboard.html',
                         username=user_id,
                         stats=stats,
                         recent_expenses=recent_expenses,
                         categories=EXPENSE_CATEGORIES)

@app.route('/expenses')
@login_required
def expenses():
    user_id = session['user_id']
    all_expenses = get_expenses(user_id)
    return render_template('expenses.html',
                         expenses=all_expenses,
                         categories=EXPENSE_CATEGORIES)

@app.route('/debts')
@login_required
def debts():
    user_id = session['user_id']
    all_debts = get_debts(user_id)
    return render_template('debts.html', debts=all_debts)

@app.route('/savings')
@login_required
def savings():
    user_id = session['user_id']
    all_savings = get_savings(user_id)
    return render_template('savings.html', savings=all_savings)

@app.route('/api/add_expense', methods=['POST'])
@login_required
def api_add_expense():
    user_id = session['user_id']
    data = request.json

    try:
        add_expense(
            user_id,
            data.get('date'),
            data.get('category'),
            data.get('description'),
            data.get('amount'),
            data.get('payment_method'),
            data.get('notes', '')
        )
        return jsonify({'success': True, 'message': 'Expense added successfully'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/add_debt', methods=['POST'])
@login_required
def api_add_debt():
    user_id = session['user_id']
    data = request.json

    try:
        add_debt(
            user_id,
            data.get('date'),
            data.get('creditor'),
            data.get('type'),
            data.get('amount'),
            data.get('status'),
            data.get('due_date'),
            data.get('notes', '')
        )
        return jsonify({'success': True, 'message': 'Debt added successfully'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/add_saving', methods=['POST'])
@login_required
def api_add_saving():
    user_id = session['user_id']
    data = request.json

    try:
        add_saving(
            user_id,
            data.get('date'),
            data.get('type'),
            data.get('description'),
            data.get('amount'),
            data.get('goal'),
            data.get('notes', '')
        )
        return jsonify({'success': True, 'message': 'Saving added successfully'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/stats')
@login_required
def api_stats():
    user_id = session['user_id']
    stats = get_summary_stats(user_id)
    return jsonify(stats)

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
